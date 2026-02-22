import streamlit as st
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import io
import numpy as np

NOTE_URL = "https://note.com/eager_roses9006/n/n935856e4df9d"

st.set_page_config(
    page_title="Image to Excel Pixel Art",
    page_icon="🎨",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    .stButton>button {
        width:100%;
        background:linear-gradient(90deg,#667eea 0%,#764ba2 100%);
        color:white; border:none;
        padding:0.75rem; font-weight:bold; border-radius:8px;
    }
    .premium-box {
        background: #fff3cd;
        border: 2px solid #f0c040;
        border-radius: 10px;
        padding: 0.8rem;
        margin: 0.4rem 0;
        color: #5a4000;
    }
    .free-box {
        background: #d4edda;
        border: 2px solid #66bb6a;
        border-radius: 10px;
        padding: 0.8rem;
        margin: 0.4rem 0;
        color: #1a4a1a;
    }
    .footer {
        text-align:center; color:#999;
        padding:2rem 0; margin-top:3rem;
        border-top:1px solid #eee; font-size:0.85rem;
    }
    input[type="password"]::-webkit-credentials-auto-fill-button,
    input[type="password"]::-webkit-strong-password-auto-fill-button {
        display: none !important;
        visibility: hidden !important;
        pointer-events: none !important;
    }
</style>
""", unsafe_allow_html=True)

# ========== サイドバー ==========
with st.sidebar:
    st.markdown("### 🎨 About This Tool")
    st.markdown("""
**Image to Excel Pixel Art Generator**
画像をExcelピクセルアートに変換

---

#### 📋 Features
- Drag & drop image upload
- Automatic image resizing
- RGB to Excel cell conversion
- Instant download

#### 使い方
1. 画像をアップロード
2. プレビューを確認
3. Excelファイルをダウンロード
""")

    st.markdown("---")

    # unlock判定（Settingsより先に処理）
    unlock_code = st.text_input(
        "🔓 解除コード / Unlock Code",
        type="password",
        placeholder="購入者はここに入力",
        autocomplete="off",
        key="unlock_code"
    )

    SECRET_CODE = "buncho"

    if unlock_code == SECRET_CODE:
        max_limit = 500
        is_premium = True
    elif unlock_code != "":
        max_limit = 100
        is_premium = False
    else:
        max_limit = 100
        is_premium = False

    st.markdown("---")

    # ========== Settings ==========
    st.markdown("### ⚙️ Settings")

    max_size = st.slider(
        "最大画像サイズ (Max Size)",
        min_value=20, max_value=max_limit,
        value=min(100, max_limit), step=10,
        help="長辺の最大ピクセル数 / 大きいほど高解像度・処理時間増"
    )
    if max_size > 200:
        st.warning(f"⚠️ {max_size}px は処理に数分かかります")

    cell_size = st.slider(
        "セルサイズ (Cell Size)",
        min_value=8, max_value=15, value=10, step=1,
        help="Excelセルのサイズ / Excel cell size in pixels"
    )

    st.markdown("---")

    # ========== Premium Unlock ==========
    st.markdown("### 🔓 Premium Unlock")
    st.markdown("""
<div class="free-box">
    <b>🆓 無料版</b><br>最大 <b>100px</b> まで使えます
</div>
<div class="premium-box">
    <b>💎 プレミアム版</b><br>最大 <b>500px</b> の高解像度<br>
    <small>解除コードを入力してください</small>
</div>
""", unsafe_allow_html=True)

    if is_premium:
        st.success("✅ プレミアム解除済み！\n500pxまで使えます🎉")
    elif unlock_code != "":
        st.error("❌ コードが違います")
    else:
        st.markdown(f"""
<div style="background:#f8f9fa;border-radius:8px;padding:0.8rem;margin:0.4rem 0;color:#333;font-size:0.9rem;">
    💎 500pxにしたい方は<br>
    <a href="{NOTE_URL}" target="_blank" style="color:#b8860b;font-weight:bold;">
    こちらのnote記事</a><br>
    で解除コードを販売中！
</div>
""", unsafe_allow_html=True)

    st.markdown("---")

    st.markdown("### 📢 おすすめ書籍")
    st.markdown("""
<div style="
    background:#fff8f0;
    border:1px solid #f0c080;
    border-radius:10px;
    padding:0.8rem;
    text-align:center;
    color:#5a3e00;
    font-size:0.88rem;
">
    🎨 <b>デザイン・イラスト</b><br>
    ピクセルアート制作に役立つ書籍<br><br>
    <a href="https://amzn.asia/d/02vCmEwo" target="_blank"
       style="
           display:inline-block;
           background:#ff9900;
           color:white;
           padding:0.4rem 1rem;
           border-radius:6px;
           text-decoration:none;
           font-weight:bold;
           font-size:0.85rem;
       ">
        🛒 Amazonで見る
    </a>
</div>
""", unsafe_allow_html=True)

    st.markdown("---")

    st.markdown("### 👤 Creator")
    st.markdown("""
**Developed by:** iueorl2

**License:** iueorl2 © 2026 Your Name. All rights reserved.
""")

# ========== メインコンテンツ ==========
st.markdown("<h1 style='text-align:center;background:linear-gradient(90deg,#667eea,#764ba2);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;font-weight:bold;'>🎨 Image to Excel Pixel Art Generator</h1>", unsafe_allow_html=True)
st.markdown("<p style='text-align:center;color:#666;'>画像をExcelピクセルアートに変換するツール<br>Transform your images into Excel pixel art</p>", unsafe_allow_html=True)

st.markdown("---")

uploaded_file = st.file_uploader(
    "画像をアップロード / Upload Image",
    type=["jpg", "jpeg", "png"],
    help="JPGまたはPNG形式の画像ファイルを選択してください"
)

def resize_image(image, max_size):
    w, h = image.size
    if w > h:
        return image.resize((max_size, int(h*(max_size/w))), Image.Resampling.LANCZOS)
    else:
        return image.resize((int(w*(max_size/h)), max_size), Image.Resampling.LANCZOS)

def create_excel_pixel_art(image, cell_size):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pixel Art"
    arr = np.array(image)
    h, w = arr.shape[:2]
    pb = st.progress(0)
    st2 = st.empty()
    for y in range(h):
        for x in range(w):
            r, g, b = arr[y, x][:3]
            hex_color = f"{int(r):02x}{int(g):02x}{int(b):02x}"
            ws.cell(row=y+1, column=x+1).fill = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )
        pb.progress((y+1)/h)
        st2.text(f"Processing... {int((y+1)/h*100)}%")
    st2.text("Adjusting cell sizes...")
    for c in range(1, w+1):
        ws.column_dimensions[get_column_letter(c)].width = cell_size / 7
    for r in range(1, h+1):
        ws.row_dimensions[r].height = cell_size
    pb.empty()
    st2.empty()
    return wb

if uploaded_file is not None:
    image = Image.open(uploaded_file)
    if image.mode != "RGB":
        image = image.convert("RGB")

    st.markdown("---")
    st.markdown("### 📸 Original Image / 元の画像")
    c1, c2 = st.columns(2)
    with c1:
        st.image(image, caption=f"Original: {image.size[0]}x{image.size[1]}px", use_container_width=True)
    resized = resize_image(image, max_size)
    with c2:
        st.image(resized, caption=f"Resized: {resized.size[0]}x{resized.size[1]}px", use_container_width=True)

    st.markdown("---")
    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        fname = st.text_input("📝 ファイル名 / File Name", value="pixel_art", help="拡張子不要")
        clean = fname.strip().replace(".xlsx", "") or "pixel_art"
        outfile = f"{clean}.xlsx"
        st.caption(f"保存ファイル名: `{outfile}`")

        if st.button("🎨 Generate Excel Pixel Art / Excelピクセルアートを生成", type="primary"):
            with st.spinner("Converting... / 変換中..."):
                try:
                    wb = create_excel_pixel_art(resized, cell_size)
                    buf = io.BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    st.success("✅ 変換完了！")
                    st.balloons()
                    st.markdown("---")
                    st.markdown("### 💾 Download Your Pixel Art")
                    c1, c2, c3 = st.columns([1, 2, 1])
                    with c2:
                        st.download_button(
                            label="📥 Download Excel File / Excelファイルをダウンロード",
                            data=buf,
                            file_name=outfile,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
else:
    st.info("👆 画像をアップロードして開始してください")

st.markdown("""
<div class="footer">
    <p>Made with ❤️ using Streamlit & Python</p>
    <p>iueorl2 © 2026 Image to Excel Pixel Art Generator. All rights reserved.</p>
</div>
""", unsafe_allow_html=True)
